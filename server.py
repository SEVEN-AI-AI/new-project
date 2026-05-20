"""
美客多广告管理系统 - Flask后端
提供API接口 + 静态文件服务 + SQLite数据库
"""

import json
import sqlite3
import os
import re
import calendar
import tempfile
from datetime import datetime
from flask import Flask, jsonify, request, send_file, send_from_directory
import openpyxl

# ── 配置 ──────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "ad_data.db")
EXCEL_DIR = BASE_DIR

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False  # 支持中文JSON响应


# ── 数据库工具 ────────────────────────────────────────────
def get_db():
    """获取数据库连接"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """初始化数据库表"""
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ad_daily_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            date TEXT NOT NULL,
            year_month TEXT NOT NULL,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            cpc_usd REAL DEFAULT 0,
            budget_usd REAL DEFAULT 0,
            organic_sales INTEGER DEFAULT 0,
            ad_sales INTEGER DEFAULT 0,
            spend REAL DEFAULT 0,
            revenue REAL DEFAULT 0,
            problem TEXT DEFAULT '',
            action TEXT DEFAULT '',
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(sku, date)
        );

        CREATE TABLE IF NOT EXISTS ad_weekly_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            year_month TEXT NOT NULL,
            week_num INTEGER NOT NULL,
            week_start TEXT NOT NULL,
            week_end TEXT NOT NULL,
            weekly_sales INTEGER DEFAULT 0,
            weekly_budget REAL DEFAULT 0,
            weekly_revenue REAL DEFAULT 0,
            weekly_acos REAL DEFAULT 0,
            keywords_ranking TEXT DEFAULT '',
            weekly_comments TEXT DEFAULT '',
            UNIQUE(sku, year_month, week_num)
        );

        CREATE TABLE IF NOT EXISTS ad_months (
            year_month TEXT PRIMARY KEY,
            label TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS ad_products (
            product_model TEXT PRIMARY KEY,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS ad_profit_daily (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            date TEXT NOT NULL,
            year_month TEXT NOT NULL,
            unit_profit REAL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(sku, date)
        );

        CREATE TABLE IF NOT EXISTS ad_table_columns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_model TEXT NOT NULL,
            year_month TEXT NOT NULL,
            col_index INTEGER NOT NULL,
            header TEXT NOT NULL,
            UNIQUE(product_model, year_month, col_index)
        );

        CREATE TABLE IF NOT EXISTS ad_table_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_model TEXT NOT NULL,
            year_month TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            date_norm TEXT DEFAULT '',
            values_json TEXT NOT NULL DEFAULT '{}',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(product_model, year_month, row_index)
        );

        CREATE INDEX IF NOT EXISTS idx_daily_sku_month
            ON ad_daily_data(sku, year_month);
        CREATE INDEX IF NOT EXISTS idx_daily_date
            ON ad_daily_data(date);
        CREATE INDEX IF NOT EXISTS idx_weekly_sku_month
            ON ad_weekly_summary(sku, year_month);
        CREATE INDEX IF NOT EXISTS idx_ad_table_rows_product_month
            ON ad_table_rows(product_model, year_month);
    """)

    conn.execute("""
        INSERT OR IGNORE INTO ad_months (year_month, label)
        SELECT DISTINCT year_month, year_month FROM ad_daily_data
    """)
    conn.commit()
    conn.close()
    print(f"[DB] 数据库已初始化: {DB_PATH}")


# ── 自然周计算 ────────────────────────────────────────────
def get_natural_weeks(year, month):
    """
    每月1号起每7天一周，最后不足7天独立成周
    例：5月 → [1-3], [4-10], [11-17], [18-24], [25-31]
    """
    days_in_month = calendar.monthrange(year, month)[1]
    first_weekday = datetime(year, month, 1).weekday()  # Monday=0, Sunday=6
    weeks = []
    day = 1
    week_num = 1
    while day <= days_in_month:
        if day == 1:
            end = min(7 - first_weekday, days_in_month)
        else:
            end = min(day + 6, days_in_month)
        weeks.append({
            'week_num': week_num,
            'start': f'{year}-{month:02d}-{day:02d}',
            'end': f'{year}-{month:02d}-{end:02d}',
        })
        day = end + 1
        week_num += 1
    return weeks


def calculate_weekly_summary(conn, sku, year_month):
    """计算并写入某SKU某月的周汇总，保留用户已输入的字段"""
    parts = year_month.split('-')
    year, month = int(parts[0]), int(parts[1])
    weeks = get_natural_weeks(year, month)

    for w in weeks:
        # 查询该周范围内的每日数据
        rows = conn.execute("""
            SELECT impressions, clicks, cpc_usd, budget_usd,
                   organic_sales, ad_sales, spend, revenue
            FROM ad_daily_data
            WHERE sku = ? AND date >= ? AND date <= ?
            ORDER BY date
        """, (sku, w['start'], w['end'])).fetchall()

        if not rows:
            continue

        weekly_sales = sum(r['organic_sales'] + r['ad_sales'] for r in rows)
        weekly_budget = sum(r['spend'] for r in rows)
        weekly_revenue = sum(r['revenue'] for r in rows)
        weekly_acos = round(weekly_budget / weekly_revenue, 4) if weekly_revenue > 0 else 0

        # 查看是否已有用户输入的字段
        existing = conn.execute("""
            SELECT keywords_ranking, weekly_comments
            FROM ad_weekly_summary
            WHERE sku = ? AND year_month = ? AND week_num = ?
        """, (sku, year_month, w['week_num'])).fetchone()

        keywords = existing['keywords_ranking'] if existing else ''
        comments = existing['weekly_comments'] if existing else ''

        conn.execute("""
            INSERT OR REPLACE INTO ad_weekly_summary
                (sku, year_month, week_num, week_start, week_end,
                 weekly_sales, weekly_budget, weekly_revenue, weekly_acos,
                 keywords_ranking, weekly_comments)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (sku, year_month, w['week_num'], w['start'], w['end'],
              weekly_sales, weekly_budget, weekly_revenue, weekly_acos,
              keywords, comments))

    conn.commit()


# ── 日期格式转换工具 ─────────────────────────────────────
def normalize_date(date_str):
    """
    将各种日期格式转为 YYYY-MM-DD
    支持: 2026.5.12, 2026-05-12, 2026/05/12, 2026.05.12
    """
    if not date_str:
        return None
    s = str(date_str).strip()
    # 尝试多种分隔符
    for sep in ['.', '/', '-']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                    # 处理浮点数情况（如 2026.0, 5.0, 12.0）
                    return f'{y:04d}-{m:02d}-{d:02d}'
                except (ValueError, IndexError):
                    return None
    return None


def parse_month_from_filename(filepath):
    """Read YYYY-MM from an ad Excel filename such as 2026.05广告数据.xlsx."""
    basename = os.path.basename(filepath)
    match = re.search(r'(\d{4})[.\-_/](\d{1,2})', basename)
    if not match:
        match = re.search(r'(\d{4})(\d{2})', basename)
    if not match:
        return datetime.now().strftime('%Y-%m')
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def cell_text(value):
    """Keep imported table cells close to the user's Excel display values."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_excel_date(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    text = cell_text(value)
    match = re.search(r'(\d{4})[./\-](\d{1,2})[./\-](\d{1,2})', text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return text


def find_ad_header_row(ws):
    """Pick the row that looks most like the ad table header."""
    best_row = 1
    best_score = -1
    max_scan_row = min(ws.max_row, 15)
    max_scan_col = min(ws.max_column, 80)
    header_keywords = [
        '日期', '曝光', '点击', '点击率', '转化率', 'CPC', '预算',
        '自然销量', '广告销量', '花费', '销售额', 'ACOS', '问题', '调整',
        '运营感受', 'date', 'spend', 'revenue',
    ]

    for row_idx in range(1, max_scan_row + 1):
        values = [cell_text(ws.cell(row_idx, col).value) for col in range(1, max_scan_col + 1)]
        non_empty = sum(1 for value in values if value)
        joined = '|'.join(values).lower()
        score = non_empty + sum(8 for keyword in header_keywords if keyword.lower() in joined)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


def import_ad_table_excel(filepath, display_name=None):
    """Import historical ad Excel as product-model monthly original tables."""
    if not os.path.exists(filepath):
        raise FileNotFoundError(f'文件不存在: {filepath}')

    import_name = display_name or filepath
    year_month = parse_month_from_filename(import_name)
    workbook = openpyxl.load_workbook(filepath, data_only=True)
    conn = get_db()
    imported_rows = 0
    products = []

    try:
        conn.execute(
            "INSERT OR IGNORE INTO ad_months (year_month, label) VALUES (?, ?)",
            (year_month, year_month),
        )

        for ws in workbook.worksheets:
            product_model = ws.title.strip()
            if not product_model:
                continue

            header_row = find_ad_header_row(ws)
            max_col = ws.max_column
            headers = []
            for col in range(1, max_col + 1):
                header = cell_text(ws.cell(header_row, col).value)
                headers.append(header or f'列{col}')

            if not any(header and not header.startswith('列') for header in headers):
                continue

            conn.execute(
                "INSERT OR IGNORE INTO ad_products (product_model) VALUES (?)",
                (product_model,),
            )
            products.append(product_model)

            conn.execute(
                "DELETE FROM ad_table_columns WHERE product_model = ? AND year_month = ?",
                (product_model, year_month),
            )
            conn.execute(
                "DELETE FROM ad_table_rows WHERE product_model = ? AND year_month = ?",
                (product_model, year_month),
            )

            for index, header in enumerate(headers, start=1):
                conn.execute(
                    """
                    INSERT INTO ad_table_columns (product_model, year_month, col_index, header)
                    VALUES (?, ?, ?, ?)
                    """,
                    (product_model, year_month, index, header),
                )

            for row_idx in range(header_row + 1, ws.max_row + 1):
                values = {
                    str(col): cell_text(ws.cell(row_idx, col).value)
                    for col in range(1, max_col + 1)
                }
                if not any(values.values()):
                    continue
                conn.execute(
                    """
                    INSERT INTO ad_table_rows
                        (product_model, year_month, row_index, date_norm, values_json)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        product_model,
                        year_month,
                        row_idx - header_row,
                        normalize_excel_date(ws.cell(row_idx, 1).value),
                        json.dumps(values, ensure_ascii=False),
                    ),
                )
                imported_rows += 1

        conn.commit()
    finally:
        conn.close()
        workbook.close()

    return {
        'ok': True,
        'file': os.path.basename(import_name),
        'month': year_month,
        'products': products,
        'rows': imported_rows,
    }


AD_TABLE_DEFAULT_HEADERS = [
    '日期', '曝光', '点击次数', '点击率', '转化率', 'CPC(USD)', 'CPC(RMB)',
    '预算(USD)', '自然销量', '广告销量', '实际广告花费', '销售额', 'ACOS',
    '问题', '调整动作', '运营感受记录', '周-销量', '周-广告预算', '周-销售额',
    '周-ACOS', '关键词-排名(周）', '周-评论',
]


def ensure_ad_product_month(conn, product_model, year_month):
    """Create a blank monthly ad table for a new product model."""
    conn.execute(
        "INSERT OR IGNORE INTO ad_products (product_model) VALUES (?)",
        (product_model,),
    )
    conn.execute(
        "INSERT OR IGNORE INTO ad_months (year_month, label) VALUES (?, ?)",
        (year_month, year_month),
    )

    existing_columns = conn.execute(
        """
        SELECT COUNT(*) AS count FROM ad_table_columns
        WHERE product_model = ? AND year_month = ?
        """,
        (product_model, year_month),
    ).fetchone()['count']
    if existing_columns:
        return

    template_rows = conn.execute(
        """
        SELECT col_index, header FROM ad_table_columns
        WHERE year_month = ?
        ORDER BY product_model, col_index
        """,
        (year_month,),
    ).fetchall()
    if not template_rows:
        template_rows = conn.execute(
            """
            SELECT col_index, header FROM ad_table_columns
            ORDER BY year_month DESC, product_model, col_index
            """
        ).fetchall()

    seen = set()
    headers = []
    for row in template_rows:
        index = row['col_index']
        if index in seen:
            continue
        seen.add(index)
        headers.append((index, row['header']))
    if not headers:
        headers = list(enumerate(AD_TABLE_DEFAULT_HEADERS, start=1))

    for index, header in headers:
        conn.execute(
            """
            INSERT INTO ad_table_columns (product_model, year_month, col_index, header)
            VALUES (?, ?, ?, ?)
            """,
            (product_model, year_month, index, header),
        )

    year, month = [int(part) for part in year_month.split('-')]
    days = calendar.monthrange(year, month)[1]
    for day in range(1, days + 1):
        date_text = f'{year}.{month}.{day}'
        date_norm = f'{year:04d}-{month:02d}-{day:02d}'
        values = {str(index): '' for index, _ in headers}
        values['1'] = date_text
        conn.execute(
            """
            INSERT INTO ad_table_rows (product_model, year_month, row_index, date_norm, values_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            (product_model, year_month, day, date_norm, json.dumps(values, ensure_ascii=False)),
        )


def number_value(value):
    text = str(value or '').strip().replace(',', '').replace('%', '')
    if not text:
        return 0
    try:
        return float(text)
    except ValueError:
        return 0


def header_matches(header, names):
    lower = str(header or '').lower()
    return any(name.lower() in lower for name in names)


# ── 静态文件路由 ──────────────────────────────────────────
@app.route('/<path:filename>')
def serve_static(filename):
    """服务项目目录下的静态文件（JS/CSS/HTML等）"""
    filepath = os.path.join(BASE_DIR, filename)
    if os.path.isfile(filepath):
        return send_from_directory(BASE_DIR, filename)
    return jsonify({'error': '文件不存在'}), 404


# ── 首页路由 ──────────────────────────────────────────────
@app.route('/')
def index():
    return send_file(os.path.join(BASE_DIR, 'erp.html'))


@app.route('/logistics')
def logistics_page():
    """物流系统独立入口"""
    return send_file(os.path.join(BASE_DIR, 'index.html'))


@app.route('/ads')
def ads_page():
    """Ads module entry; use the unified ERP first-level menu."""
    return send_file(os.path.join(BASE_DIR, 'erp.html'))


@app.route('/ads-table')
def ads_table_page():
    """Legacy raw ad-detail page is retired; keep the URL returning ERP."""
    return send_file(os.path.join(BASE_DIR, 'erp.html'))


@app.route('/api/ads', methods=['GET'])
def get_ads():
    """获取每日广告数据，支持 ?sku=D001&month=2026-05 筛选"""
    sku = request.args.get('sku', '')
    month = request.args.get('month', '')

    conn = get_db()
    query = "SELECT * FROM ad_daily_data WHERE 1=1"
    params = []

    if sku:
        query += " AND sku = ?"
        params.append(sku)
    if month:
        query += " AND year_month = ?"
        params.append(month)

    query += " ORDER BY sku, date"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    result = [dict(r) for r in rows]
    return jsonify(result)


@app.route('/api/ads/<int:ad_id>', methods=['PUT'])
def update_ad(ad_id):
    """更新单条每日数据"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '无数据'}), 400

    conn = get_db()

    # 检查记录是否存在
    row = conn.execute("SELECT * FROM ad_daily_data WHERE id = ?", (ad_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '记录不存在'}), 404

    # 更新字段
    fields = ['impressions', 'clicks', 'cpc_usd', 'budget_usd',
              'organic_sales', 'ad_sales', 'spend', 'revenue',
              'problem', 'action', 'note']
    sets = []
    values = []
    for f in fields:
        if f in data:
            sets.append(f"{f} = ?")
            values.append(data[f])

    if sets:
        sets.append("updated_at = CURRENT_TIMESTAMP")
        values.append(ad_id)
        conn.execute(
            f"UPDATE ad_daily_data SET {', '.join(sets)} WHERE id = ?",
            values
        )
        # 更新后重算该月的周汇总
        conn.commit()
        calculate_weekly_summary(conn, row['sku'], row['year_month'])

    conn.close()
    return jsonify({'ok': True})


@app.route('/api/ads/<int:ad_id>', methods=['DELETE'])
def delete_ad(ad_id):
    """删除单条每日数据"""
    conn = get_db()
    row = conn.execute("SELECT * FROM ad_daily_data WHERE id = ?", (ad_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': '记录不存在'}), 404

    sku = row['sku']
    month = row['year_month']
    conn.execute("DELETE FROM ad_daily_data WHERE id = ?", (ad_id,))
    conn.commit()
    calculate_weekly_summary(conn, sku, month)
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/ads/bulk', methods=['POST'])
def bulk_create_ads():
    """批量写入每日数据（脚本同步用）"""
    data = request.get_json()
    if not data or 'records' not in data:
        return jsonify({'error': '需要 records 数组'}), 400

    conn = get_db()
    inserted = 0
    updated = 0
    affected_months = set()

    for rec in data['records']:
        date_norm = normalize_date(rec.get('date', ''))
        if not date_norm:
            continue

        sku = rec.get('sku', '')
        year_month = date_norm[:7]  # YYYY-MM

        affected_months.add((sku, year_month))

        conn.execute("""
            INSERT OR REPLACE INTO ad_daily_data
                (sku, date, year_month, impressions, clicks, cpc_usd,
                 budget_usd, organic_sales, ad_sales, spend, revenue,
                 problem, action, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            sku, date_norm, year_month,
            rec.get('impressions', 0),
            rec.get('clicks', 0),
            rec.get('cpc_usd', 0),
            rec.get('budget_usd', 0),
            rec.get('organic_sales', 0),
            rec.get('ad_sales', 0),
            rec.get('spend', 0),
            rec.get('revenue', 0),
            rec.get('problem', ''),
            rec.get('action', ''),
            rec.get('note', ''),
        ))
        inserted += 1

    conn.commit()

    # 重算受影响的月份周汇总
    for sku, month in affected_months:
        calculate_weekly_summary(conn, sku, month)

    conn.close()
    return jsonify({'ok': True, 'count': inserted})


@app.route('/api/ads/skus', methods=['GET'])
def get_skus():
    """Return SKU-like product models, including user-created models with no imported rows yet."""
    conn = get_db()
    rows = conn.execute("""
        SELECT sku FROM (
            SELECT DISTINCT sku FROM ad_daily_data WHERE sku <> ''
            UNION
            SELECT product_model AS sku FROM ad_products WHERE product_model <> ''
        )
        ORDER BY sku
    """).fetchall()
    conn.close()
    skus = [r['sku'] for r in rows if re.match(r'^[A-Za-z0-9][A-Za-z0-9_-]{1,}$', r['sku'] or '')]
    return jsonify(skus)


@app.route('/api/ads/skus', methods=['POST'])
def create_sku():
    data = request.get_json() or {}
    sku = str(data.get('sku', '')).strip()
    if not sku:
        return jsonify({'error': '\u8bf7\u586b\u5199\u578b\u53f7'}), 400
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO ad_products (product_model) VALUES (?)", (sku,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'sku': sku})


@app.route('/api/ads/skus/<path:sku>', methods=['DELETE'])
def delete_sku(sku):
    sku = sku.strip()
    if not sku:
        return jsonify({'error': '\u578b\u53f7\u4e0d\u80fd\u4e3a\u7a7a'}), 400
    conn = get_db()
    conn.execute("DELETE FROM ad_daily_data WHERE sku = ?", (sku,))
    conn.execute("DELETE FROM ad_weekly_summary WHERE sku = ?", (sku,))
    conn.execute("DELETE FROM ad_profit_daily WHERE sku = ?", (sku,))
    conn.execute("DELETE FROM ad_products WHERE product_model = ?", (sku,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'sku': sku})


@app.route('/api/profit', methods=['GET'])
def get_profit_rows():
    sku = request.args.get('sku', '').strip()
    month = request.args.get('month', '').strip()
    if not sku or not month:
        return jsonify([])

    conn = get_db()
    rows = conn.execute("""
        SELECT d.sku, d.date, d.year_month,
               SUM(COALESCE(d.organic_sales, 0) + COALESCE(d.ad_sales, 0)) AS quantity,
               SUM(COALESCE(d.spend, 0)) AS ad_spend,
               COALESCE(MAX(p.unit_profit), 0) AS unit_profit
        FROM ad_daily_data d
        LEFT JOIN ad_profit_daily p ON p.sku = d.sku AND p.date = d.date
        WHERE d.sku = ? AND d.year_month = ?
        GROUP BY d.sku, d.date, d.year_month
        ORDER BY d.date
    """, (sku, month)).fetchall()
    conn.close()

    result = []
    for row in rows:
        quantity = int(row['quantity'] or 0)
        unit_profit = float(row['unit_profit'] or 0)
        ad_spend = float(row['ad_spend'] or 0)
        gross_profit = quantity * unit_profit
        net_profit = gross_profit - ad_spend
        result.append({
            'sku': row['sku'],
            'date': row['date'],
            'year_month': row['year_month'],
            'quantity': quantity,
            'unit_profit': unit_profit,
            'gross_profit': gross_profit,
            'ad_spend': ad_spend,
            'net_profit': net_profit,
        })
    return jsonify(result)


@app.route('/api/profit', methods=['PUT'])
def update_profit_row():
    data = request.get_json() or {}
    sku = str(data.get('sku', '')).strip()
    date_norm = normalize_date(data.get('date', ''))
    if not sku or not date_norm:
        return jsonify({'error': '\u7f3a\u5c11\u578b\u53f7\u6216\u65e5\u671f'}), 400
    unit_profit = float(data.get('unit_profit') or 0)
    conn = get_db()
    conn.execute("""
        INSERT INTO ad_profit_daily (sku, date, year_month, unit_profit, updated_at)
        VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(sku, date) DO UPDATE SET
            unit_profit = excluded.unit_profit,
            updated_at = CURRENT_TIMESTAMP
    """, (sku, date_norm, date_norm[:7], unit_profit))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/ads/months', methods=['GET'])
def get_months():
    """Get all month tabs, including empty months created by users."""
    conn = get_db()
    rows = conn.execute("""
        SELECT year_month FROM ad_months
        UNION
        SELECT DISTINCT year_month FROM ad_daily_data
        ORDER BY year_month
    """).fetchall()
    conn.close()
    return jsonify([r['year_month'] for r in rows])


@app.route('/api/ads/months', methods=['POST'])
def create_month():
    """Create an empty month tab."""
    data = request.get_json() or {}
    year_month = str(data.get('month', '')).strip()
    if not re.match(r'^\d{4}-\d{2}$', year_month):
        return jsonify({'error': 'Month format must be YYYY-MM'}), 400

    year, month = map(int, year_month.split('-'))
    if month < 1 or month > 12:
        return jsonify({'error': 'Month must be between 01 and 12'}), 400

    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO ad_months (year_month, label) VALUES (?, ?)",
        (year_month, year_month)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'month': year_month})


@app.route('/api/ads/months/<year_month>', methods=['DELETE'])
def delete_month(year_month):
    """Delete a month tab and the related daily and weekly data."""
    if not re.match(r'^\d{4}-\d{2}$', year_month):
        return jsonify({'error': 'Month format must be YYYY-MM'}), 400

    conn = get_db()
    conn.execute("DELETE FROM ad_daily_data WHERE year_month = ?", (year_month,))
    conn.execute("DELETE FROM ad_weekly_summary WHERE year_month = ?", (year_month,))
    conn.execute("DELETE FROM ad_months WHERE year_month = ?", (year_month,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'month': year_month})

@app.route('/api/weekly', methods=['GET'])
def get_weekly():
    """获取周汇总，支持 ?sku=D001&month=2026-05 筛选"""
    sku = request.args.get('sku', '')
    month = request.args.get('month', '')

    conn = get_db()
    query = "SELECT * FROM ad_weekly_summary WHERE 1=1"
    params = []

    if sku:
        query += " AND sku = ?"
        params.append(sku)
    if month:
        query += " AND year_month = ?"
        params.append(month)

    query += " ORDER BY sku, year_month, week_num"

    rows = conn.execute(query, params).fetchall()
    conn.close()

    result = [dict(r) for r in rows]
    return jsonify(result)


@app.route('/api/weekly/<int:week_id>', methods=['PUT'])
def update_weekly(week_id):
    """更新周汇总（关键词排名/周评论）"""
    data = request.get_json()
    if not data:
        return jsonify({'error': '无数据'}), 400

    conn = get_db()
    fields = []
    values = []

    if 'keywords_ranking' in data:
        fields.append("keywords_ranking = ?")
        values.append(data['keywords_ranking'])
    if 'weekly_comments' in data:
        fields.append("weekly_comments = ?")
        values.append(data['weekly_comments'])

    if fields:
        values.append(week_id)
        conn.execute(
            f"UPDATE ad_weekly_summary SET {', '.join(fields)} WHERE id = ?",
            values
        )
        conn.commit()

    conn.close()
    return jsonify({'ok': True})


@app.route('/api/weekly/calculate', methods=['POST'])
def calculate_weekly():
    """重新计算某月的周汇总"""
    data = request.get_json()
    month = data.get('month', '') if data else ''

    conn = get_db()
    if month:
        skus = conn.execute(
            "SELECT DISTINCT sku FROM ad_daily_data WHERE year_month = ?",
            (month,)
        ).fetchall()
        for row in skus:
            calculate_weekly_summary(conn, row['sku'], month)
    else:
        # 重算所有月份
        months = conn.execute(
            "SELECT DISTINCT year_month FROM ad_daily_data"
        ).fetchall()
        for m in months:
            skus = conn.execute(
                "SELECT DISTINCT sku FROM ad_daily_data WHERE year_month = ?",
                (m['year_month'],)
            ).fetchall()
            for s in skus:
                calculate_weekly_summary(conn, s['sku'], m['year_month'])

    conn.close()
    return jsonify({'ok': True})


# ── API: Excel同步/导出 ──────────────────────────────────

def sync_excel_file(filepath, display_name=None):
    """Import ad Excel data from a server path or uploaded temp file."""
    if not os.path.exists(filepath):
        return {'error': f'?????: {filepath}'}, 400

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {'error': f'??Excel??: {e}'}, 400

    basename = display_name or os.path.basename(filepath)
    month_match = re.search(r'(\d{4})[.\-](\d{1,2})', basename)
    if not month_match:
        wb.close()
        return {'error': f'??????????: {basename}'}, 400

    year_month = f"{month_match.group(1)}-{int(month_match.group(2)):02d}"
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO ad_months (year_month, label) VALUES (?, ?)", (year_month, year_month))
    records = []
    sku_list = ['D001', 'D005', 'D007', '048', '091', 'PM002']

    for sku in sku_list:
        if sku not in wb.sheetnames:
            continue
        ws = wb[sku]
        for r in range(3, ws.max_row + 1):
            date_norm = normalize_date(ws.cell(r, 1).value)
            if not date_norm or not date_norm.startswith(year_month):
                continue
            impressions = ws.cell(r, 2).value
            clicks = ws.cell(r, 3).value
            if not impressions and not clicks:
                continue
            records.append({
                'sku': sku,
                'date': date_norm,
                'year_month': year_month,
                'impressions': int(impressions or 0),
                'clicks': int(clicks or 0),
                'cpc_usd': float(ws.cell(r, 6).value or 0),
                'budget_usd': float(ws.cell(r, 8).value or 0),
                'organic_sales': int(ws.cell(r, 9).value or 0),
                'ad_sales': int(ws.cell(r, 10).value or 0),
                'spend': float(ws.cell(r, 11).value or 0),
                'revenue': float(ws.cell(r, 12).value or 0),
                'problem': str(ws.cell(r, 14).value or '').strip(),
                'action': str(ws.cell(r, 15).value or '').strip(),
                'note': str(ws.cell(r, 16).value or '').strip(),
            })

    for rec in records:
        conn.execute("INSERT OR IGNORE INTO ad_products (product_model) VALUES (?)", (rec['sku'],))
        conn.execute('''
            INSERT OR REPLACE INTO ad_daily_data
                (sku, date, year_month, impressions, clicks, cpc_usd,
                 budget_usd, organic_sales, ad_sales, spend, revenue,
                 problem, action, note)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            rec['sku'], rec['date'], rec['year_month'], rec['impressions'], rec['clicks'],
            rec['cpc_usd'], rec['budget_usd'], rec['organic_sales'], rec['ad_sales'],
            rec['spend'], rec['revenue'], rec['problem'], rec['action'], rec['note']
        ))

    conn.commit()
    affected_skus = sorted({rec['sku'] for rec in records})
    for sku in affected_skus:
        calculate_weekly_summary(conn, sku, year_month)
    conn.close()
    wb.close()
    return {'ok': True, 'file': basename, 'month': year_month, 'records': len(records), 'skus': affected_skus}, 200


@app.route('/api/sync/excel', methods=['POST'])
def sync_excel():
    """???????????Excel???"""
    data = request.get_json() or {}
    filepath = str(data.get('path', '')).strip()
    if not filepath:
        return jsonify({'error': '?? path ??'}), 400
    payload, status = sync_excel_file(filepath)
    return jsonify(payload), status


@app.route('/api/sync/excel-file', methods=['POST'])
def sync_excel_upload():
    """???????Excel???????"""
    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'error': '??? Excel ??'}), 400
    if not upload.filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
        return jsonify({'error': '??? Excel ??'}), 400

    suffix = os.path.splitext(upload.filename)[1] or '.xlsx'
    temp_path = ''
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp_path = temp.name
            upload.save(temp)
        payload, status = sync_excel_file(temp_path, upload.filename)
        return jsonify(payload), status
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """导出指定月份数据为Excel"""
    month = request.args.get('month', '')
    sku = request.args.get('sku', '')

    conn = get_db()
    query = "SELECT * FROM ad_daily_data WHERE 1=1"
    params = []

    if month:
        query += " AND year_month = ?"
        params.append(month)
    if sku:
        query += " AND sku = ?"
        params.append(sku)

    query += " ORDER BY sku, date"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    if not rows:
        return jsonify({'error': '无数据可导出'}), 404

    # 创建Excel
    wb = openpyxl.Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 按SKU分组写入
    sku_data = {}
    for r in rows:
        d = dict(r)
        s = d['sku']
        if s not in sku_data:
            sku_data[s] = []
        sku_data[s].append(d)

    for sku_name, data_rows in sku_data.items():
        ws = wb.create_sheet(title=sku_name)

        # 表头
        headers = ['日期', '曝光', '点击', '点击率', '转化率', 'CPC(USD)',
                   'CPC(RMB)', '预算', '自然销量', '广告销量', '花费',
                   '销售额', 'ACOS', '问题', '调整动作', '运营感受']
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)

        for i, row in enumerate(data_rows, 2):
            ws.cell(i, 1, row['date'])
            ws.cell(i, 2, row['impressions'])
            ws.cell(i, 3, row['clicks'])
            # 点击率 = 公式
            ws.cell(i, 4, f'=IFERROR(C{i}/B{i},0)')
            # 转化率 = 公式
            ws.cell(i, 5, f'=IFERROR(J{i}/C{i},0)')
            ws.cell(i, 6, row['cpc_usd'])
            # CPC(RMB) = 公式
            ws.cell(i, 7, f'=F{i}*0.38')
            ws.cell(i, 8, row['budget_usd'])
            ws.cell(i, 9, row['organic_sales'])
            ws.cell(i, 10, row['ad_sales'])
            ws.cell(i, 11, row['spend'])
            ws.cell(i, 12, row['revenue'])
            # ACOS = 公式
            ws.cell(i, 13, f'=IFERROR(K{i}/L{i},0)')
            ws.cell(i, 14, row['problem'])
            ws.cell(i, 15, row['action'])
            ws.cell(i, 16, row['note'])

    # 保存到临时文件
    filename = f"广告数据导出_{month or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filename = f"ad_export_{month or 'all'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    filepath = os.path.join(EXCEL_DIR, filename)
    wb.save(filepath)
    wb.close()

    return send_file(filepath, as_attachment=True, download_name=filename)


# ── 启动 ──────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    init_db()
    print("=" * 60)
    print("Mercado Libre Ads backend started")
    print("ERP: http://localhost:5000")
    print("Ads: http://localhost:5000/ads")
    print("DB:", DB_PATH)
    print("=" * 60)
    app.run(host='0.0.0.0', port=5000, debug=True)
