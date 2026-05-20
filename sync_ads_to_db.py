"""
美客多广告历史数据 → SQLite 同步脚本
扫描Excel目录下所有月度xlsx文件，导入到SQLite数据库
"""

import os
import re
import sqlite3
import openpyxl

# ── 配置 ──────────────────────────────────────────────────
DB_PATH = r"D:\平台\美客多-跨境\广告\ad_data.db"
EXCEL_DIR = r"D:\平台\美客多-跨境\广告"
SKU_LIST = ['D001', 'D005', 'D007', '048', '091', 'PM002']


def normalize_date(date_str):
    """将各种日期格式转为 YYYY-MM-DD"""
    if not date_str:
        return None
    s = str(date_str).strip()
    for sep in ['.', '/', '-']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    y, m, d = int(float(parts[0])), int(float(parts[1])), int(float(parts[2]))
                    return f'{y:04d}-{m:02d}-{d:02d}'
                except (ValueError, IndexError):
                    return None
    return None


def init_db():
    """确保数据库表存在"""
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ad_daily_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            date TEXT NOT NULL,
            year_month TEXT NOT NULL,
            impressions INTEGER DEFAULT 0,
            clicks INTEGER DEFAULT 0,
            cpc_usd REAL DEFAULT 0,
            budget_usd REAL DEFAULT 0,
            organic_sales INTEGER DEFAULT 0,
            ad_sales INTEGER DEFAULT 0,
            spend REAL DEFAULT 0,
            revenue REAL DEFAULT 0,
            problem TEXT DEFAULT '',
            action TEXT DEFAULT '',
            note TEXT DEFAULT '',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(sku, date)
        );
        CREATE TABLE IF NOT EXISTS ad_weekly_summary (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT NOT NULL,
            year_month TEXT NOT NULL,
            week_num INTEGER NOT NULL,
            week_start TEXT NOT NULL,
            week_end TEXT NOT NULL,
            weekly_sales INTEGER DEFAULT 0,
            weekly_budget REAL DEFAULT 0,
            weekly_revenue REAL DEFAULT 0,
            weekly_acos REAL DEFAULT 0,
            keywords_ranking TEXT DEFAULT '',
            weekly_comments TEXT DEFAULT '',
            UNIQUE(sku, year_month, week_num)
        );
    """)
    conn.commit()
    conn.close()


def calculate_weekly(conn, sku, year_month):
    """重算某SKU某月的周汇总"""
    import calendar
    parts = year_month.split('-')
    year, month = int(parts[0]), int(parts[1])
    days_in_month = calendar.monthrange(year, month)[1]

    weeks = []
    day = 1
    week_num = 1
    while day <= days_in_month:
        end = min(day + 6, days_in_month)
        weeks.append((week_num, f'{year:04d}-{month:02d}-{day:02d}', f'{year:04d}-{month:02d}-{end:02d}'))
        day = end + 1
        week_num += 1

    for wn, ws, we in weeks:
        rows = conn.execute("""
            SELECT organic_sales, ad_sales, spend, revenue
            FROM ad_daily_data
            WHERE sku = ? AND date >= ? AND date <= ?
        """, (sku, ws, we)).fetchall()

        if not rows:
            continue

        w_sales = sum(r['organic_sales'] + r['ad_sales'] for r in rows)
        w_budget = sum(r['spend'] for r in rows)
        w_revenue = sum(r['revenue'] for r in rows)
        w_acos = round(w_budget / w_revenue, 4) if w_revenue > 0 else 0

        existing = conn.execute("""
            SELECT keywords_ranking, weekly_comments
            FROM ad_weekly_summary WHERE sku=? AND year_month=? AND week_num=?
        """, (sku, year_month, wn)).fetchone()

        conn.execute("""
            INSERT OR REPLACE INTO ad_weekly_summary
                (sku, year_month, week_num, week_start, week_end,
                 weekly_sales, weekly_budget, weekly_revenue, weekly_acos,
                 keywords_ranking, weekly_comments)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (sku, year_month, wn, ws, we, w_sales, w_budget, w_revenue, w_acos,
              existing['keywords_ranking'] if existing else '',
              existing['weekly_comments'] if existing else ''))

    conn.commit()


def sync_one_file(filepath):
    """同步单个Excel文件"""
    basename = os.path.basename(filepath)

    # 从文件名提取月份
    month_match = re.search(r'(\d{4})[.\-](\d{1,2})', basename)
    if not month_match:
        print(f"  跳过（无法识别月份）: {basename}")
        return 0

    year_month = f"{month_match.group(1)}-{int(month_match.group(2)):02d}"
    print(f"  文件: {basename} → 月份: {year_month}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  打开失败: {e}")
        return 0

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    count = 0

    for sku in SKU_LIST:
        if sku not in wb.sheetnames:
            print(f"    Sheet '{sku}' 不存在，跳过")
            continue

        ws = wb[sku]
        sku_count = 0

        for r in range(3, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if not date_val:
                continue

            date_norm = normalize_date(date_val)
            if not date_norm:
                continue

            if not date_norm.startswith(year_month):
                continue

            impressions = ws.cell(r, 2).value
            clicks = ws.cell(r, 3).value
            cpc_usd = ws.cell(r, 6).value
            budget_usd = ws.cell(r, 8).value
            organic_sales = ws.cell(r, 9).value
            ad_sales = ws.cell(r, 10).value
            spend = ws.cell(r, 11).value
            revenue = ws.cell(r, 12).value

            problem = ws.cell(r, 14).value or ''
            action = ws.cell(r, 15).value or ''
            note = ws.cell(r, 16).value or ''

            if not impressions and not clicks:
                continue

            conn.execute("""
                INSERT OR REPLACE INTO ad_daily_data
                    (sku, date, year_month, impressions, clicks, cpc_usd,
                     budget_usd, organic_sales, ad_sales, spend, revenue,
                     problem, action, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                sku, date_norm, year_month,
                int(impressions or 0), int(clicks or 0),
                float(cpc_usd or 0), float(budget_usd or 0),
                int(organic_sales or 0), int(ad_sales or 0),
                float(spend or 0), float(revenue or 0),
                str(problem).strip(), str(action).strip(), str(note).strip()
            ))
            sku_count += 1

        if sku_count:
            print(f"    SKU {sku}: {sku_count} 条")
            calculate_weekly(conn, sku, year_month)

        count += sku_count

    conn.commit()
    conn.close()
    wb.close()
    return count


def main():
    print("=" * 55)
    print("  美客多广告历史数据 → SQLite 同步")
    print("=" * 55)

    init_db()

    # 扫描所有xlsx文件
    xlsx_files = []
    for f in os.listdir(EXCEL_DIR):
        if f.endswith('.xlsx') and 'backup' not in f.lower():
            xlsx_files.append(os.path.join(EXCEL_DIR, f))

    xlsx_files.sort()
    print(f"\n找到 {len(xlsx_files)} 个Excel文件:\n")

    total = 0
    for fp in xlsx_files:
        total += sync_one_file(fp)

    print(f"\n{'=' * 55}")
    print(f"  同步完成！共导入 {total} 条数据")
    print(f"  数据库: {DB_PATH}")
    print(f"{'=' * 55}")


if __name__ == '__main__':
    main()
